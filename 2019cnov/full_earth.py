import time, json, requests
import pandas as pd
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
url = 'https://view.inews.qq.com/g2/getOnsInfo?name=disease_h5&callback=&_=%d'%int(time.time()*1000)
data = json.loads(requests.get(url=url).json()['data'])
print(data.keys())
d = data['areaTree']
print(data)
name=list()           #国家
now_confirm=[]
confirm_list = list()  # 确诊
suspect_list = list()  # 疑似
dead_list = list()  # 死亡
heal_list = list()  # 治愈
for item in d:
    name.append(item['name'])
    confirm_list.append(int(item['total']['confirm']))
    suspect_list.append(int(item['total']['suspect']))
    dead_list.append(int(item['total']['dead']))
    heal_list.append(int(item['total']['heal']))
print(len(confirm_list))
now_confirm=np.array(confirm_list)-np.array(dead_list)-np.array(heal_list)
print(confirm_list)
now=time.strftime("%m-%d-%Y")
workbook = openpyxl.load_workbook(now+'.xlsx')
earth = workbook.create_sheet(title='earth',index=1)
earth.cell(row=1, column=1, value='国家')
earth.cell(row=1, column=2, value='确诊')
earth.cell(row=1, column=3, value='疑似')
earth.cell(row=1, column=4, value='死亡')
earth.cell(row=1, column=5, value='治愈')
earth.cell(row=1, column=6, value='现存确诊')
for i in range(1,len(name)):
    earth.cell(row=i+1, column=1, value=name[i-1])
for i in range(1, len(name)):
    earth.cell(row=i+1, column=2, value=confirm_list[i - 1])
for i in range(1, len(name)):
    earth.cell(row=i+1, column=3, value=suspect_list[i - 1])
for i in range(1, len(name)):
    earth.cell(row=i+1, column=4, value=dead_list[i - 1])
for i in range(1, len(name)):
    earth.cell(row=i+1, column=5, value=heal_list[i - 1])
for i in range(1, len(name)):
    earth.cell(row=i+1, column=6, value=now_confirm[i - 1])
workbook.save(now+'.xlsx')#保存文件
